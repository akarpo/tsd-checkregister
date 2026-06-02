"""Rebuild the master Troy_SD_Check_Register workbook from parsed all_lines.pkl.

Produces all 7 sheets matching the existing schema.
Uses openpyxl + os.replace pattern to be SharePoint-sync friendly.
"""
from __future__ import annotations
import pickle, os, tempfile
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BUILD = Path(__file__).parent
ALL_LINES = pickle.loads((BUILD / 'all_lines.pkl').read_bytes())

REPO = Path(r'C:\Dev\CheckRegister')
OUT_NAME = 'Troy_SD_Check_Register_FY20-FY26.xlsx'
INTERMEDIATE = REPO / '_staging.xlsx'
FINAL = REPO / OUT_NAME

# FY range
fys_present = sorted({r['Fiscal Year'] for r in ALL_LINES})
print(f'FYs present: {fys_present}')
print(f'Total rows: {len(ALL_LINES):,}')

wb = openpyxl.Workbook()

# =================== Sheet 1: README ===================
ws = wb.active
ws.title = 'README'
readme = [
    'Troy School District — Combined Check Register — FY' + fys_present[0][2:] + ' through ' + fys_present[-1] + '-to-date',
    'Source: BoardDocs (https://go.boarddocs.com/mi/troysd/Board.nsf/Public)',
    f'Records: {len(ALL_LINES):,} line items',
    f"Total amount: =SUM('All Lines'!P:P)",
    'Sheets:',
    '  All Lines       — every parsed transaction. Filterable by fund, vendor, budget unit, etc.',
    '  By Year x Fund  — total spend per fund per fiscal year',
    '  By Budget Unit  — totals per budget unit, with FY-by-FY breakdown',
    '  Curriculum_PD   — every line item charged to budget unit 101-425-221 (curriculum & PD)',
    '  PD_Yearly_Summary — Curriculum/PD spend by fiscal year',
    '  PD by Subject   — Curriculum/PD spending classified by subject area',
    'Notes:',
    '  - "Function Code" is positions 7-9 of the budget unit (Michigan PSAM function code).',
    '    111-119 = Direct instruction; 21x = Pupil Services; 22x = Instructional Staff (PD/Curriculum);',
    '    25x = Central Services; 26x = Operations & Maintenance; 27x = Transportation; 28x = Other Support;',
    '    33x = Athletics & Activities; 45x = Capital Outlay; 51x = Debt Service.',
    '  - "Account" is the object code (3xxx = purchased services, 4xxx = supplies, 5xxx = capital,',
    '    L4xxx = liability/payroll, etc.)',
    '  - "Voided" = "V" indicates a voided check (amount is negative).',
    '  - Some lines have a 3-digit budget_unit (e.g. just "101") — these are payroll deductions',
    '    or benefit liability transfers (account codes starting with L), not departmental spending.',
    '  - First two registers (Jul/Aug 2022 meetings) cover FY22 activity; included for completeness.',
    'Subject classification (Curriculum_PD / PD by Subject):',
    '  - Rule-based using vendor names and (often-truncated) description keywords.',
    '  - Confidence: "high" = vendor or strong keyword match; "med" = vendor inferred from primary business; "low" = no subject signal.',
    '  - Not Directly Attributable means a generic "consulting" or "PD" line where the description does not name a subject.',
]
for txt in readme:
    ws.append([txt])
ws.column_dimensions['A'].width = 95

# =================== Sheet 2: All Lines ===================
ws = wb.create_sheet('All Lines')
header = ['Source Meeting','Fiscal Year','Fund','Fund Name','Cash Acct','Check No','Voided',
         'Issue Date','Vendor ID','Vendor Name','Budget Unit','Function Code','Account',
         'Description','Sales Tax','Amount']
ws.append(header)
for cell in ws[1]:
    cell.font = Font(bold=True)
for r in ALL_LINES:
    ws.append([r.get(h) for h in header])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
# Column widths
widths = {'A':14,'B':6,'C':6,'D':24,'E':10,'F':12,'G':7,'H':12,'I':10,'J':22,'K':18,'L':9,'M':9,'N':22,'O':10,'P':14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
print(f'  All Lines: {len(ALL_LINES):,} rows')

# =================== Sheet 3: By Year x Fund ===================
ws = wb.create_sheet('By Year x Fund')
ws.append(['Fiscal Year','Fund','Fund Name','Line Count','Total Amount'])
for cell in ws[1]: cell.font = Font(bold=True)
agg = defaultdict(lambda: {'count': 0, 'amt': 0.0})
fund_names = {}
for r in ALL_LINES:
    k = (r['Fiscal Year'], r['Fund'], r['Fund Name'])
    agg[k]['count'] += 1
    agg[k]['amt'] += r['Amount']
    fund_names[r['Fund']] = r['Fund Name']
for (fy, fund, fname), v in sorted(agg.items()):
    ws.append([fy, fund, fname, v['count'], v['amt']])
ws.freeze_panes = 'A2'
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 16
print(f'  By Year x Fund: {len(agg)} rows')

# =================== Sheet 4: By Budget Unit ===================
ws = wb.create_sheet('By Budget Unit')
fy_list = sorted(fys_present)
ws.append(['Budget Unit','Function Code'] + [f'{fy} Lines' for fy in fy_list] + [f'{fy} Amount' for fy in fy_list] + ['Total Lines','Total Amount'])
for cell in ws[1]: cell.font = Font(bold=True)
bu_agg = defaultdict(lambda: {fy: {'c': 0, 'a': 0.0} for fy in fy_list})
bu_func = {}
for r in ALL_LINES:
    bu = r['Budget Unit']
    fy = r['Fiscal Year']
    bu_agg[bu][fy]['c'] += 1
    bu_agg[bu][fy]['a'] += r['Amount']
    bu_func[bu] = r['Function Code']
# Sort by total amount desc
items = sorted(bu_agg.items(), key=lambda x: -sum(x[1][fy]['a'] for fy in fy_list))
for bu, byfy in items:
    counts = [byfy[fy]['c'] for fy in fy_list]
    amts = [byfy[fy]['a'] for fy in fy_list]
    ws.append([bu, bu_func[bu]] + counts + amts + [sum(counts), sum(amts)])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
print(f'  By Budget Unit: {len(items)} rows')

# =================== Sheet 5: PD_Yearly_Summary ===================
ws = wb.create_sheet('PD_Yearly_Summary')
ws.append(['Fiscal Year','Line Count','Total Spend'])
for cell in ws[1]: cell.font = Font(bold=True)
pd_lines = [r for r in ALL_LINES if r['Budget Unit'].startswith('101425221')]
fy_pd = defaultdict(lambda: {'c': 0, 'a': 0.0})
for r in pd_lines:
    fy_pd[r['Fiscal Year']]['c'] += 1
    fy_pd[r['Fiscal Year']]['a'] += r['Amount']
for fy in sorted(fy_pd):
    ws.append([fy, fy_pd[fy]['c'], fy_pd[fy]['a']])
ws.append(['Total', sum(v['c'] for v in fy_pd.values()), sum(v['a'] for v in fy_pd.values())])
print(f'  PD_Yearly_Summary: {len(fy_pd)} FYs, {len(pd_lines):,} PD lines')

# =================== Sheet 6: Curriculum_PD ===================
ws = wb.create_sheet('Curriculum_PD')
ws.append(['Filter: Budget Unit starts with "101425221" (Function 221 = Improvement of Instruction)'])
ws.append([f'Records: {len(pd_lines):,}    Total: =SUM(K4:K{len(pd_lines)+3})'])
ws.append(['Source Meeting','Fiscal Year','Fund','Issue Date','Vendor Name','Budget Unit','Account','Description','Subject','Confidence','Amount'])
for cell in ws[3]: cell.font = Font(bold=True)
for r in pd_lines:
    ws.append([r['Source Meeting'], r['Fiscal Year'], r['Fund'], r['Issue Date'], r['Vendor Name'],
              r['Budget Unit'], r['Account'], r['Description'], r['Subject'], r['Confidence'], r['Amount']])
ws.freeze_panes = 'A4'
print(f'  Curriculum_PD: {len(pd_lines):,} rows')

# =================== Sheet 7: PD by Subject ===================
ws = wb.create_sheet('PD by Subject')
ws.append(['Curriculum / PD spend (Budget Unit 101-425-221) — classified by subject area'])
ws.append(['Classification: rule-based, on vendor name + truncated description text. Mark "Confidence" col on Curriculum_PD sheet for spot-checks.'])
ws.append([])
header = ['Subject','Lines'] + [f'{fy} $' for fy in fy_list] + ['Total $']
ws.append(header)
for cell in ws[4]: cell.font = Font(bold=True)
subj_agg = defaultdict(lambda: {'c': 0, **{fy: 0.0 for fy in fy_list}})
for r in pd_lines:
    s = r['Subject'] or 'Not Directly Attributable'
    subj_agg[s]['c'] += 1
    subj_agg[s][r['Fiscal Year']] += r['Amount']
items = sorted(subj_agg.items(), key=lambda x: -sum(x[1][fy] for fy in fy_list))
for subj, v in items:
    row = [subj, v['c']] + [v[fy] for fy in fy_list]
    row.append(sum(v[fy] for fy in fy_list))
    ws.append(row)
total_row = ['TOTAL', sum(v['c'] for v in subj_agg.values())] + [sum(v[fy] for v in subj_agg.values()) for fy in fy_list]
total_row.append(sum(total_row[2:]))
ws.append(total_row)
print(f'  PD by Subject: {len(items)} subjects')

# =================== Save ===================
wb.save(INTERMEDIATE)
os.replace(INTERMEDIATE, FINAL)
print(f'\nSaved: {FINAL}')
print(f'Size: {FINAL.stat().st_size:,} bytes')
