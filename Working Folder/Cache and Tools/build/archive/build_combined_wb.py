"""Apply categorization + subject classification to combined dataset, build master FY12-FY26 workbook."""
import pickle, sys, os
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent))
from categorize_v2 import categorize, VENDOR_CATS

combined = pickle.loads(Path(r'C:\Dev\CheckRegister\Working Folder\Cache and Tools\build\combined_lines.pkl').read_bytes())
print(f'Loaded: {len(combined):,} rows', flush=True)

VENDOR_SUBJECT = pickle.loads(Path(r'C:\Users\Alex\AppData\Local\Temp\vendor_subject.pkl').read_bytes())

def classify_subject(vendor, desc):
    if not vendor:
        return 'Not Directly Attributable'
    for v_key, subj in VENDOR_SUBJECT.items():
        if vendor.startswith(v_key[:15]):
            return subj
    d = (desc or '').upper()
    if any(k in d for k in ('READING', 'WRITING', 'LITERACY', 'CALKINS', 'F&P', 'FOUNTAS')):
        return 'ELA'
    if any(k in d for k in ('MATH', 'AVMR', 'MRSP', 'BRIDGES', 'ALGEBRA', 'GEOMETRY')):
        return 'Math'
    if 'SCIENCE' in d:
        return 'Science'
    if 'SOCIAL' in d:
        return 'Social Studies'
    return 'Not Directly Attributable'

print('Applying categorization...', flush=True)
for r in combined:
    r['Category'] = categorize(r.get('Vendor Name', ''), r.get('Fund', ''),
                               r.get('Function Code', ''), r.get('Account', ''),
                               r.get('Budget Unit', ''), r.get('Amount', 0))
    if r.get('Budget Unit', '').startswith('101425221'):
        r['Subject'] = classify_subject(r.get('Vendor Name', ''), r.get('Description', ''))
        v = r.get('Vendor Name', '')
        if v and any(v.startswith(vk[:15]) for vk in VENDOR_SUBJECT):
            r['Confidence'] = 'high'
        elif r.get('Subject') != 'Not Directly Attributable':
            r['Confidence'] = 'med'
        else:
            r['Confidence'] = 'low'
    else:
        r['Subject'] = ''
        r['Confidence'] = ''

Path(r'C:\Dev\CheckRegister\Working Folder\Cache and Tools\build\combined_lines.pkl').write_bytes(pickle.dumps(combined))
print('Saved updated combined_lines.pkl', flush=True)

print('Building workbook...', flush=True)
wb = openpyxl.Workbook()
fys_present = sorted({r['Fiscal Year'] for r in combined})
total_amt = sum(r['Amount'] for r in combined)

ws = wb.active
ws.title = 'README'
readme = [
    f'Troy School District — Combined Check Register — {fys_present[0]} through {fys_present[-1]}',
    'Source: BoardDocs (https://go.boarddocs.com/mi/troysd/Board.nsf/Public)',
    f'Records: {len(combined):,} line items',
    f'Total amount: ${total_amt:,.2f}',
    '',
    'COVERAGE NOTES:',
    '  FY12-FY19 data extracted from embedded check registers in monthly meeting-packet PDFs',
    '  FY20 missing - no separate registers in 2019-09 to 2020-07 BoardDocs (transition period)',
    '  FY21+ data extracted from standalone "Check register by fund" PDFs (BoardDocs began separating Aug 2020)',
    '',
    'Sheets:',
    '  All Lines       - every parsed transaction',
    '  By Year x Fund  - total spend per fund per fiscal year',
    '  By Budget Unit  - totals per budget unit, with FY breakdown',
    '  Curriculum_PD   - every line charged to budget unit 101-425-221 (curriculum & PD)',
    '  PD_Yearly_Summary - Curriculum/PD spend by fiscal year',
    '  PD by Subject   - Curriculum/PD spending classified by subject area',
]
for txt in readme:
    ws.append([txt])
ws.column_dimensions['A'].width = 95

ws = wb.create_sheet('All Lines')
header = ['Source Meeting', 'Fiscal Year', 'Fund', 'Fund Name', 'Cash Acct', 'Check No', 'Voided',
          'Issue Date', 'Vendor ID', 'Vendor Name', 'Budget Unit', 'Function Code', 'Account',
          'Description', 'Sales Tax', 'Amount']
ws.append(header)
for cell in ws[1]:
    cell.font = Font(bold=True)
for r in combined:
    ws.append([r.get(h) for h in header])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
print(f'  All Lines: {len(combined):,} rows written', flush=True)

ws = wb.create_sheet('By Year x Fund')
ws.append(['Fiscal Year', 'Fund', 'Fund Name', 'Line Count', 'Total Amount'])
for cell in ws[1]:
    cell.font = Font(bold=True)
agg = defaultdict(lambda: {'count': 0, 'amt': 0.0})
for r in combined:
    k = (r['Fiscal Year'], r['Fund'], r['Fund Name'])
    agg[k]['count'] += 1
    agg[k]['amt'] += r['Amount']
for k, v in sorted(agg.items()):
    ws.append([k[0], k[1], k[2], v['count'], v['amt']])
ws.freeze_panes = 'A2'

ws = wb.create_sheet('By Budget Unit')
fy_list = fys_present
ws.append(['Budget Unit', 'Function Code'] + [f'{fy} Lines' for fy in fy_list] + [f'{fy} Amount' for fy in fy_list] + ['Total Lines', 'Total Amount'])
for cell in ws[1]:
    cell.font = Font(bold=True)
bu_agg = defaultdict(lambda: {fy: {'c': 0, 'a': 0.0} for fy in fy_list})
bu_func = {}
for r in combined:
    bu = r['Budget Unit']
    fy = r['Fiscal Year']
    bu_agg[bu][fy]['c'] += 1
    bu_agg[bu][fy]['a'] += r['Amount']
    bu_func[bu] = r['Function Code']
items = sorted(bu_agg.items(), key=lambda x: -sum(x[1][fy]['a'] for fy in fy_list))
for bu, byfy in items:
    counts = [byfy[fy]['c'] for fy in fy_list]
    amts = [byfy[fy]['a'] for fy in fy_list]
    ws.append([bu, bu_func[bu]] + counts + amts + [sum(counts), sum(amts)])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
print(f'  By Budget Unit: {len(items)} rows', flush=True)

ws = wb.create_sheet('PD_Yearly_Summary')
ws.append(['Fiscal Year', 'Line Count', 'Total Spend'])
for cell in ws[1]:
    cell.font = Font(bold=True)
pd_lines = [r for r in combined if r['Budget Unit'].startswith('101425221')]
fy_pd = defaultdict(lambda: {'c': 0, 'a': 0.0})
for r in pd_lines:
    fy_pd[r['Fiscal Year']]['c'] += 1
    fy_pd[r['Fiscal Year']]['a'] += r['Amount']
for fy in sorted(fy_pd):
    ws.append([fy, fy_pd[fy]['c'], fy_pd[fy]['a']])
ws.append(['Total', sum(v['c'] for v in fy_pd.values()), sum(v['a'] for v in fy_pd.values())])
print(f'  PD lines: {len(pd_lines):,}', flush=True)

ws = wb.create_sheet('Curriculum_PD')
ws.append(['Filter: Budget Unit starts with "101425221" (Function 221 = Improvement of Instruction)'])
ws.append([f'Records: {len(pd_lines):,}'])
ws.append(['Source Meeting', 'Fiscal Year', 'Fund', 'Issue Date', 'Vendor Name', 'Budget Unit', 'Account', 'Description', 'Subject', 'Confidence', 'Amount'])
for cell in ws[3]:
    cell.font = Font(bold=True)
for r in pd_lines:
    ws.append([r['Source Meeting'], r['Fiscal Year'], r['Fund'], r['Issue Date'], r['Vendor Name'],
               r['Budget Unit'], r['Account'], r['Description'], r['Subject'], r['Confidence'], r['Amount']])
ws.freeze_panes = 'A4'

ws = wb.create_sheet('PD by Subject')
ws.append(['Curriculum / PD spend (Budget Unit 101-425-221) - classified by subject area'])
ws.append([])
ws.append(['Subject', 'Lines'] + [f'{fy} $' for fy in fy_list] + ['Total $'])
for cell in ws[3]:
    cell.font = Font(bold=True)
subj_agg = defaultdict(lambda: {'c': 0, **{fy: 0.0 for fy in fy_list}})
for r in pd_lines:
    s = r['Subject'] or 'Not Directly Attributable'
    subj_agg[s]['c'] += 1
    subj_agg[s][r['Fiscal Year']] += r['Amount']
items = sorted(subj_agg.items(), key=lambda x: -sum(x[1][fy] for fy in fy_list))
for subj, v in items:
    row = [subj, v['c']] + [v[fy] for fy in fy_list]
    row.append(sum(v[fy] for fy in fy_list))
    ws.append(row)
total_row = ['TOTAL', sum(v['c'] for v in subj_agg.values())] + [sum(v[fy] for v in subj_agg.values()) for fy in fy_list]
total_row.append(sum(total_row[2:]))
ws.append(total_row)

INTERMEDIATE = Path(r'C:\Dev\CheckRegister\_staging.xlsx')
FINAL = Path(r'C:\Dev\CheckRegister\Troy_SD_Check_Register_FY12-FY26.xlsx')
wb.save(INTERMEDIATE)
os.replace(INTERMEDIATE, FINAL)
print(f'\nSaved: {FINAL}', flush=True)
print(f'Size: {FINAL.stat().st_size:,} bytes', flush=True)
