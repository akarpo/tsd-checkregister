"""Rebuild combined dataset:
1. Add the Oct 2019 recovered rows
2. Add an 'Issue Date FY' column (FY computed from actual transaction Issue Date)
3. Apply categorization + subject (re-run for all rows incl. recovered)
4. Rebuild master workbook with summary sheets pivoted on BOTH meeting FY and Issue-Date FY
"""
import sys, pickle, os
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent))
from categorize_v2 import categorize, VENDOR_CATS

BUILD = Path(__file__).parent
combined = pickle.loads((BUILD / 'combined_lines.pkl').read_bytes())
recovered = pickle.loads((BUILD / 'oct2019_recovered.pkl').read_bytes())
print(f'Existing combined: {len(combined):,}', flush=True)
print(f'Adding recovered Oct 2019: {len(recovered):,}', flush=True)

# Merge — recovered fits chronologically between FY19 and FY21 (it's FY20 by meeting; mostly FY20 by issue)
# Insert at the right position: after FY19 rows
combined.extend(recovered)
print(f'Total: {len(combined):,}', flush=True)

# Add Issue Date FY column
def issue_fy(d):
    if not isinstance(d, datetime):
        return ''
    return f'FY{(d.year+1)%100:02d}' if d.month >= 7 else f'FY{d.year%100:02d}'

VENDOR_SUBJECT = pickle.loads(Path(r'C:\Users\Alex\AppData\Local\Temp\vendor_subject.pkl').read_bytes())

def classify_subject(vendor, desc):
    if not vendor: return 'Not Directly Attributable'
    for v_key, subj in VENDOR_SUBJECT.items():
        if vendor.startswith(v_key[:15]):
            return subj
    d = (desc or '').upper()
    if any(k in d for k in ('READING', 'WRITING', 'LITERACY', 'CALKINS', 'F&P', 'FOUNTAS')):
        return 'ELA'
    if any(k in d for k in ('MATH', 'AVMR', 'MRSP', 'BRIDGES', 'ALGEBRA', 'GEOMETRY')):
        return 'Math'
    if 'SCIENCE' in d: return 'Science'
    if 'SOCIAL' in d: return 'Social Studies'
    return 'Not Directly Attributable'

print('Re-applying categorization + subjects + Issue-Date FY...', flush=True)
for r in combined:
    r['Issue Date FY'] = issue_fy(r['Issue Date'])
    r['Category'] = categorize(r.get('Vendor Name',''), r.get('Fund',''),
                              r.get('Function Code',''), r.get('Account',''),
                              r.get('Budget Unit',''), r.get('Amount', 0))
    if r.get('Budget Unit','').startswith('101425221'):
        r['Subject'] = classify_subject(r.get('Vendor Name',''), r.get('Description',''))
        v = r.get('Vendor Name','')
        if v and any(v.startswith(vk[:15]) for vk in VENDOR_SUBJECT):
            r['Confidence'] = 'high'
        elif r.get('Subject') != 'Not Directly Attributable':
            r['Confidence'] = 'med'
        else:
            r['Confidence'] = 'low'
    else:
        r['Subject'] = ''
        r['Confidence'] = ''

(BUILD / 'combined_lines.pkl').write_bytes(pickle.dumps(combined))
print('Saved updated combined_lines.pkl', flush=True)

# Build the workbook with both FY pivots
print('Building workbook...', flush=True)
wb = openpyxl.Workbook()
mfys = sorted({r['Fiscal Year'] for r in combined})
ifys = sorted({r['Issue Date FY'] for r in combined if r['Issue Date FY']})
total_amt = sum(r['Amount'] for r in combined)

# 1. README
ws = wb.active
ws.title = 'README'
readme = [
    f'Troy School District - Combined Check Register - FY12 through FY26',
    'Source: BoardDocs (https://go.boarddocs.com/mi/troysd/Board.nsf/Public)',
    f'Records: {len(combined):,} line items',
    f'Total amount: ${total_amt:,.2f}',
    '',
    'TWO FISCAL-YEAR COLUMNS:',
    '  Fiscal Year       = FY of the BOARD MEETING that approved the register',
    '  Issue Date FY     = FY of the actual TRANSACTION DATE (true FY)',
    '  These differ because a register typically covers 1-2 months of prior activity.',
    '  For accurate FY-by-FY analysis, use Issue Date FY.',
    '',
    'COVERAGE NOTES:',
    f'  FY12-FY19 from embedded check registers in monthly meeting packets (PARTIAL coverage)',
    f'  FY20 partially recovered: 1,418 rows from Oct 2019 meeting (recovered with pypdf after pdfplumber error)',
    f'  Pre-2020 missing months: see project_docs/INDEX.md for the complete gap list',
    f'  FY21+ from standalone "Check register by fund" PDFs',
    '',
    'Sheets:',
    '  All Lines               - every transaction (now with both Fiscal Year and Issue Date FY)',
    '  By Year x Fund          - by MEETING FY x fund',
    '  By Issue-Date FY x Fund - by TRUE TRANSACTION FY x fund (recommended for FY analysis)',
    '  By Budget Unit          - by budget unit, with FY breakdown',
    '  Curriculum_PD           - filtered to budget unit 101-425-221',
    '  PD_Yearly_Summary       - PD spend by both FY conventions',
    '  PD by Subject           - PD subject classification by year',
]
for txt in readme:
    ws.append([txt])
ws.column_dimensions['A'].width = 95

# 2. All Lines (with new Issue Date FY column)
ws = wb.create_sheet('All Lines')
header = ['Source Meeting','Fiscal Year','Issue Date FY','Fund','Fund Name','Cash Acct','Check No','Voided',
          'Issue Date','Vendor ID','Vendor Name','Budget Unit','Function Code','Account',
          'Description','Sales Tax','Amount']
ws.append(header)
for cell in ws[1]: cell.font = Font(bold=True)
for r in combined:
    ws.append([r.get(h) for h in header])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
print(f'  All Lines: {len(combined):,} rows written', flush=True)

# 3. By Year x Fund (meeting FY)
ws = wb.create_sheet('By Year x Fund')
ws.append(['Meeting FY','Fund','Fund Name','Line Count','Total Amount'])
for cell in ws[1]: cell.font = Font(bold=True)
agg = defaultdict(lambda: {'count': 0, 'amt': 0.0})
for r in combined:
    k = (r['Fiscal Year'], r['Fund'], r['Fund Name'])
    agg[k]['count'] += 1
    agg[k]['amt'] += r['Amount']
for k, v in sorted(agg.items()):
    ws.append([k[0], k[1], k[2], v['count'], v['amt']])
ws.freeze_panes = 'A2'

# 4. By Issue-Date FY x Fund (TRUE FY)
ws = wb.create_sheet('By Issue-Date FY x Fund')
ws.append(['Issue Date FY','Fund','Fund Name','Line Count','Total Amount'])
for cell in ws[1]: cell.font = Font(bold=True)
agg = defaultdict(lambda: {'count': 0, 'amt': 0.0})
for r in combined:
    if not r.get('Issue Date FY'): continue
    k = (r['Issue Date FY'], r['Fund'], r['Fund Name'])
    agg[k]['count'] += 1
    agg[k]['amt'] += r['Amount']
for k, v in sorted(agg.items()):
    ws.append([k[0], k[1], k[2], v['count'], v['amt']])
ws.freeze_panes = 'A2'

# 5. By Budget Unit (uses Meeting FY for backward-compat with original schema)
ws = wb.create_sheet('By Budget Unit')
fy_list = mfys
ws.append(['Budget Unit', 'Function Code'] + [f'{fy} Lines' for fy in fy_list] + [f'{fy} Amount' for fy in fy_list] + ['Total Lines', 'Total Amount'])
for cell in ws[1]: cell.font = Font(bold=True)
bu_agg = defaultdict(lambda: {fy: {'c': 0, 'a': 0.0} for fy in fy_list})
bu_func = {}
for r in combined:
    bu = r['Budget Unit']
    fy = r['Fiscal Year']
    bu_agg[bu][fy]['c'] += 1
    bu_agg[bu][fy]['a'] += r['Amount']
    bu_func[bu] = r['Function Code']
items = sorted(bu_agg.items(), key=lambda x: -sum(x[1][fy]['a'] for fy in fy_list))
for bu, byfy in items:
    counts = [byfy[fy]['c'] for fy in fy_list]
    amts = [byfy[fy]['a'] for fy in fy_list]
    ws.append([bu, bu_func[bu]] + counts + amts + [sum(counts), sum(amts)])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
print(f'  By Budget Unit: {len(items)} rows', flush=True)

# 6. PD Yearly Summary — both FY conventions
ws = wb.create_sheet('PD_Yearly_Summary')
ws.append(['MEETING FY (legacy)', 'Line Count', 'Total Spend', '', 'ISSUE-DATE FY (true)', 'Line Count', 'Total Spend'])
for cell in ws[1]: cell.font = Font(bold=True)
pd_lines = [r for r in combined if r['Budget Unit'].startswith('101425221')]
fy_pd_meeting = defaultdict(lambda: {'c': 0, 'a': 0.0})
fy_pd_issue = defaultdict(lambda: {'c': 0, 'a': 0.0})
for r in pd_lines:
    fy_pd_meeting[r['Fiscal Year']]['c'] += 1
    fy_pd_meeting[r['Fiscal Year']]['a'] += r['Amount']
    if r.get('Issue Date FY'):
        fy_pd_issue[r['Issue Date FY']]['c'] += 1
        fy_pd_issue[r['Issue Date FY']]['a'] += r['Amount']
all_fys = sorted(set(fy_pd_meeting) | set(fy_pd_issue))
for fy in all_fys:
    mc = fy_pd_meeting[fy]['c']
    ma = fy_pd_meeting[fy]['a']
    ic = fy_pd_issue[fy]['c']
    ia = fy_pd_issue[fy]['a']
    ws.append([fy, mc, ma, '', fy, ic, ia])
ws.append(['Total', sum(v['c'] for v in fy_pd_meeting.values()), sum(v['a'] for v in fy_pd_meeting.values()), '',
           'Total', sum(v['c'] for v in fy_pd_issue.values()), sum(v['a'] for v in fy_pd_issue.values())])

# 7. Curriculum_PD (with Issue Date FY column)
ws = wb.create_sheet('Curriculum_PD')
ws.append(['Filter: Budget Unit starts with "101425221" (Function 221 = Improvement of Instruction)'])
ws.append([f'Records: {len(pd_lines):,}'])
ws.append(['Source Meeting','Meeting FY','Issue Date FY','Fund','Issue Date','Vendor Name','Budget Unit','Account','Description','Subject','Confidence','Amount'])
for cell in ws[3]: cell.font = Font(bold=True)
for r in pd_lines:
    ws.append([r['Source Meeting'], r['Fiscal Year'], r.get('Issue Date FY',''), r['Fund'], r['Issue Date'], r['Vendor Name'],
               r['Budget Unit'], r['Account'], r['Description'], r['Subject'], r['Confidence'], r['Amount']])
ws.freeze_panes = 'A4'

# 8. PD by Subject — pivoted on Issue Date FY (TRUE FY)
ws = wb.create_sheet('PD by Subject')
ws.append(['Curriculum / PD spend (Budget Unit 101-425-221) - classified by subject area'])
ws.append(['Pivot: ISSUE DATE FY (true fiscal year of the transaction, not the approving meeting)'])
ws.append([])
fy_list = sorted(set(ifys))
ws.append(['Subject','Lines'] + [f'{fy} $' for fy in fy_list] + ['Total $'])
for cell in ws[4]: cell.font = Font(bold=True)
subj_agg = defaultdict(lambda: {'c': 0, **{fy: 0.0 for fy in fy_list}})
for r in pd_lines:
    s = r['Subject'] or 'Not Directly Attributable'
    subj_agg[s]['c'] += 1
    if r.get('Issue Date FY'):
        subj_agg[s][r['Issue Date FY']] += r['Amount']
items = sorted(subj_agg.items(), key=lambda x: -sum(x[1][fy] for fy in fy_list))
for subj, v in items:
    row = [subj, v['c']] + [v[fy] for fy in fy_list]
    row.append(sum(v[fy] for fy in fy_list))
    ws.append(row)
total_row = ['TOTAL', sum(v['c'] for v in subj_agg.values())] + [sum(v[fy] for v in subj_agg.values()) for fy in fy_list]
total_row.append(sum(total_row[2:]))
ws.append(total_row)

# Save
INTERMEDIATE = Path(r'C:\Dev\CheckRegister\_staging.xlsx')
FINAL = Path(r'C:\Dev\CheckRegister\Troy_SD_Check_Register_FY12-FY26.xlsx')
wb.save(INTERMEDIATE)
os.replace(INTERMEDIATE, FINAL)
print(f'Saved: {FINAL} ({FINAL.stat().st_size:,} bytes)', flush=True)
