"""Build the master workbook (8 sheets: Meeting-FY + Issue-Date-FY pivots) from
combined, classified rows.

Exposes ``build_workbook(combined, out_path)`` so the orchestrator (rebuild.py)
and the validator (validate.py) can drive it with in-memory rows. Run directly
to (re)write the committed workbook from build/combined_lines.pkl.
"""
import sys
import os
import pickle
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _paths


def build_workbook(combined, out_path):
    """Build the 8-sheet workbook from `combined` rows; atomic-write to out_path.

    Returns (out_path, row_count, grand_total)."""
    out_path = Path(out_path)
    mfys = sorted({r['Fiscal Year'] for r in combined})
    ifys = sorted({r['Issue Date FY'] for r in combined if r.get('Issue Date FY')})
    total_amt = sum(r['Amount'] for r in combined)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'README'
    readme = [
        f'Troy School District - Combined Check Register - FY11 through FY26',
        'Source: BoardDocs (https://go.boarddocs.com/mi/troysd/Board.nsf/Public)',
        f'Records: {len(combined):,} line items',
        f'Total amount: ${total_amt:,.2f}',
        '',
        'TWO FISCAL-YEAR COLUMNS:',
        '  Fiscal Year       = FY of the BOARD MEETING that approved the register',
        '  Issue Date FY     = FY of the actual TRANSACTION DATE (true FY)',
        '  These differ because a register typically covers 1-2 months of prior activity.',
        '  For accurate FY-by-FY analysis, use Issue Date FY.',
        '',
        'COVERAGE NOTES (after bundle-date fix May 2026):',
        '  FY11-FY19 from embedded check registers in monthly meeting packets',
        '  FY20 partially recovered: Oct 2019 register via pypdf fallback + portions of FY19/FY20',
        '  FY21-FY26 from standalone "Check register by fund" PDFs',
        '',
        'Sheets:',
        '  All Lines               - every transaction with both FY columns',
        '  By Year x Fund          - by MEETING FY x fund',
        '  By Issue-Date FY x Fund - by TRUE TRANSACTION FY x fund (recommended)',
        '  By Budget Unit          - by budget unit',
        '  Curriculum_PD           - filtered to budget unit 101-425-221',
        '  PD_Yearly_Summary       - PD spend by both FY conventions',
        '  PD by Subject           - PD subject classification by Issue-Date FY',
    ]
    for txt in readme:
        ws.append([txt])
    ws.column_dimensions['A'].width = 95

    # All Lines
    ws = wb.create_sheet('All Lines')
    header = ['Source Meeting','Fiscal Year','Issue Date FY','Fund','Fund Name','Cash Acct','Check No','Voided',
              'Issue Date','Vendor ID','Vendor Name','Budget Unit','Function Code','Account',
              'Description','Sales Tax','Amount']
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in combined:
        ws.append([r.get(h) for h in header])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # By Year x Fund (meeting FY)
    ws = wb.create_sheet('By Year x Fund')
    ws.append(['Meeting FY','Fund','Fund Name','Line Count','Total Amount'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    agg = defaultdict(lambda: {'count': 0, 'amt': 0.0})
    for r in combined:
        k = (r['Fiscal Year'], r['Fund'], r['Fund Name'])
        agg[k]['count'] += 1; agg[k]['amt'] += r['Amount']
    for k, v in sorted(agg.items()):
        ws.append([k[0], k[1], k[2], v['count'], v['amt']])
    ws.freeze_panes = 'A2'

    # By Issue-Date FY x Fund
    ws = wb.create_sheet('By Issue-Date FY x Fund')
    ws.append(['Issue Date FY','Fund','Fund Name','Line Count','Total Amount'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    agg = defaultdict(lambda: {'count': 0, 'amt': 0.0})
    for r in combined:
        if not r.get('Issue Date FY'): continue
        k = (r['Issue Date FY'], r['Fund'], r['Fund Name'])
        agg[k]['count'] += 1; agg[k]['amt'] += r['Amount']
    for k, v in sorted(agg.items()):
        ws.append([k[0], k[1], k[2], v['count'], v['amt']])
    ws.freeze_panes = 'A2'

    # By Budget Unit (meeting FY for backward-compat)
    ws = wb.create_sheet('By Budget Unit')
    fy_list = mfys
    ws.append(['Budget Unit','Function Code'] + [f'{fy} Lines' for fy in fy_list] + [f'{fy} Amount' for fy in fy_list] + ['Total Lines','Total Amount'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    bu_agg = defaultdict(lambda: {fy: {'c': 0, 'a': 0.0} for fy in fy_list})
    bu_func = {}
    for r in combined:
        bu = r['Budget Unit']; fy = r['Fiscal Year']
        bu_agg[bu][fy]['c'] += 1; bu_agg[bu][fy]['a'] += r['Amount']
        bu_func[bu] = r['Function Code']
    items = sorted(bu_agg.items(), key=lambda x: -sum(x[1][fy]['a'] for fy in fy_list))
    for bu, byfy in items:
        counts = [byfy[fy]['c'] for fy in fy_list]
        amts = [byfy[fy]['a'] for fy in fy_list]
        ws.append([bu, bu_func[bu]] + counts + amts + [sum(counts), sum(amts)])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # PD_Yearly_Summary — both conventions
    ws = wb.create_sheet('PD_Yearly_Summary')
    ws.append(['MEETING FY','Line Count','Total Spend','','ISSUE-DATE FY (true)','Line Count','Total Spend'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    pd_lines = [r for r in combined if r['Budget Unit'].startswith('101425221')]
    fy_pd_meeting = defaultdict(lambda: {'c': 0, 'a': 0.0})
    fy_pd_issue = defaultdict(lambda: {'c': 0, 'a': 0.0})
    for r in pd_lines:
        fy_pd_meeting[r['Fiscal Year']]['c'] += 1; fy_pd_meeting[r['Fiscal Year']]['a'] += r['Amount']
        if r.get('Issue Date FY'):
            fy_pd_issue[r['Issue Date FY']]['c'] += 1; fy_pd_issue[r['Issue Date FY']]['a'] += r['Amount']
    all_fys = sorted(set(fy_pd_meeting) | set(fy_pd_issue))
    for fy in all_fys:
        ws.append([fy, fy_pd_meeting[fy]['c'], fy_pd_meeting[fy]['a'], '',
                   fy, fy_pd_issue[fy]['c'], fy_pd_issue[fy]['a']])
    ws.append(['Total', sum(v['c'] for v in fy_pd_meeting.values()), sum(v['a'] for v in fy_pd_meeting.values()), '',
               'Total', sum(v['c'] for v in fy_pd_issue.values()), sum(v['a'] for v in fy_pd_issue.values())])

    # Curriculum_PD
    ws = wb.create_sheet('Curriculum_PD')
    ws.append(['Filter: Budget Unit starts with "101425221" (Function 221 = Improvement of Instruction)'])
    ws.append([f'Records: {len(pd_lines):,}'])
    ws.append(['Source Meeting','Meeting FY','Issue Date FY','Fund','Issue Date','Vendor Name','Budget Unit','Account','Description','Subject','Confidence','Amount'])
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for r in pd_lines:
        ws.append([r['Source Meeting'], r['Fiscal Year'], r.get('Issue Date FY',''), r['Fund'],
                   r['Issue Date'], r['Vendor Name'], r['Budget Unit'], r['Account'],
                   r['Description'], r['Subject'], r['Confidence'], r['Amount']])
    ws.freeze_panes = 'A4'

    # PD by Subject — pivoted on Issue Date FY
    ws = wb.create_sheet('PD by Subject')
    ws.append(['Curriculum / PD spend (Budget Unit 101-425-221) - classified by subject area'])
    ws.append(['Pivot: ISSUE DATE FY (true fiscal year)'])
    ws.append([])
    fy_list = sorted(ifys)
    ws.append(['Subject','Lines'] + [f'{fy} $' for fy in fy_list] + ['Total $'])
    for cell in ws[4]:
        cell.font = Font(bold=True)
    subj_agg = defaultdict(lambda: {'c': 0, **{fy: 0.0 for fy in fy_list}})
    for r in pd_lines:
        s = r['Subject'] or 'Not Directly Attributable'
        subj_agg[s]['c'] += 1
        if r.get('Issue Date FY'):
            subj_agg[s][r['Issue Date FY']] += r['Amount']
    items = sorted(subj_agg.items(), key=lambda x: -sum(x[1][fy] for fy in fy_list))
    for subj, v in items:
        row = [subj, v['c']] + [v[fy] for fy in fy_list]
        row.append(sum(v[fy] for fy in fy_list))
        ws.append(row)
    total_row = ['TOTAL', sum(v['c'] for v in subj_agg.values())] + [sum(v[fy] for v in subj_agg.values()) for fy in fy_list]
    total_row.append(sum(total_row[2:]))
    ws.append(total_row)

    # Atomic write (staging file beside the target, then os.replace)
    staging = out_path.with_name(out_path.stem + '_staging.xlsx')
    wb.save(staging)
    os.replace(staging, out_path)
    return out_path, len(combined), total_amt


if __name__ == '__main__':
    combined = pickle.loads(_paths.COMBINED_PKL.read_bytes())
    print(f'Loaded: {len(combined):,} rows', flush=True)
    path, n, amt = build_workbook(combined, _paths.WORKBOOK)
    print(f'Saved: {path} ({Path(path).stat().st_size:,} bytes) - {n:,} rows / ${amt:,.2f}', flush=True)
