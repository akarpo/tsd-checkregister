"""Non-destructive validation of the consolidated assemble stage.

The source PDFs for a full from-scratch rebuild are mostly not in the repo, but
the *assemble* stage (rows -> workbook + dashboard) is fully checkable against
the committed deliverables. We reconstruct the combined row set from the
committed workbook's All Lines sheet (re-deriving Category / Subject /
Confidence from the committed lookups), rebuild the workbook to a scratch file
(never touching the committed one), and check:

  A. All Lines is reproduced cell-for-cell (data round-trips).
  B. Subject re-derivation reproduces every committed Curriculum_PD subject
     (proves vendor_subject.json is a faithful reconstruction).
  C. Every rebuilt pivot reconciles to the grand total (internal consistency).
  D. The rebuilt dashboard payload reconciles to the workbook.

These four are the PASS criteria. Separately, the committed workbook is a
patchwork assembled across several script versions + a partial recovery merge,
so its *summary* sheets differ from a clean rebuild; those differences are
reported as informational findings (a single rebuild.py run brings them into
sync). Run:  python validate.py
"""
import sys
import json
import re
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _paths
from categorize_v2 import categorize
from subjects import classify_subject, classify_confidence
from rebuild_final import build_workbook
from rebuild_dashboard_full import build_payload

SCRATCH = _paths.BUILD / "_validate_scratch.xlsx"
HEADER = ['Source Meeting', 'Fiscal Year', 'Issue Date FY', 'Fund', 'Fund Name', 'Cash Acct',
          'Check No', 'Voided', 'Issue Date', 'Vendor ID', 'Vendor Name', 'Budget Unit',
          'Function Code', 'Account', 'Description', 'Sales Tax', 'Amount']


def reconstruct_rows(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    it = wb["All Lines"].iter_rows(values_only=True)
    next(it)
    rows = []
    for vals in it:
        if vals[0] is None and vals[1] is None:
            continue
        r = dict(zip(HEADER, vals))
        vendor = r["Vendor Name"] or ""
        bu = str(r["Budget Unit"] or "")
        r["Category"] = categorize(vendor, str(r["Fund"] or ""), str(r["Function Code"] or ""),
                                   str(r["Account"] or ""), bu, r["Amount"] or 0)
        if bu.startswith("101425221"):
            subj = classify_subject(vendor, r["Description"] or "")
            r["Subject"], r["Confidence"] = subj, classify_confidence(vendor, subj)
        else:
            r["Subject"] = r["Confidence"] = ""
        rows.append(r)
    wb.close()
    return rows


def load_grid(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True)
    grid = [list(row) for row in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    return grid


def cell_eq(a, b, tol=0.01):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) and not isinstance(a, bool):
        return abs(a - b) <= tol
    return a == b


def count_grid_diffs(gc, gr):
    if len(gc) != len(gr):
        return None
    diffs = 0
    for rc, rr in zip(gc, gr):
        if len(rc) != len(rr):
            diffs += 1
            continue
        diffs += sum(1 for a, b in zip(rc, rr) if not cell_eq(a, b))
    return diffs


def pivot_totals(grid, line_col, amt_col, skip):
    lines = amt = 0.0
    for row in grid[skip:]:
        if line_col < len(row) and isinstance(row[line_col], (int, float)):
            lines += row[line_col]
        if amt_col < len(row) and isinstance(row[amt_col], (int, float)):
            amt += row[amt_col]
    return int(lines), amt


def main():
    print("Reconstructing rows from committed workbook All Lines...", flush=True)
    rows = reconstruct_rows(_paths.WORKBOOK)
    N = len(rows)
    GRAND = round(sum(r["Amount"] for r in rows), 2)
    print(f"  {N:,} rows / ${GRAND:,.2f}", flush=True)

    print("Rebuilding workbook to scratch file (non-destructive)...", flush=True)
    build_workbook(rows, SCRATCH)

    checks = []  # (name, passed, detail)

    # A. All Lines reproduced cell-for-cell
    diffs = count_grid_diffs(load_grid(_paths.WORKBOOK, "All Lines"), load_grid(SCRATCH, "All Lines"))
    checks.append(("All Lines reproduced cell-for-cell", diffs == 0, f"{N:,} rows, {diffs} cell diffs"))

    # B. Subject re-derivation matches committed Curriculum_PD
    gc = load_grid(_paths.WORKBOOK, "Curriculum_PD")
    data = [r for r in gc if r and r[0] not in (None, "Source Meeting")
            and not str(r[0]).startswith(("Filter:", "Records:"))]
    subj_mism = sum(1 for r in data if r[9] != classify_subject(r[5] or "", r[8] or ""))
    checks.append(("Subject lookup reproduces committed subjects", subj_mism == 0,
                   f"{len(data):,} PD lines, {subj_mism} mismatches"))

    # C. Each rebuilt pivot reconciles to the grand total
    for sheet, lc, ac, sk in [("By Year x Fund", 3, 4, 1),
                              ("By Issue-Date FY x Fund", 3, 4, 1),
                              ("By Budget Unit", -2, -1, 1)]:
        g = load_grid(SCRATCH, sheet)
        lines, amt = pivot_totals(g, lc, ac, sk)
        ok = lines == N and abs(amt - GRAND) < 0.5
        checks.append((f"Pivot reconciles: {sheet}", ok, f"{lines:,} lines / ${amt:,.2f}"))

    # D. Rebuilt dashboard payload reconciles to the workbook
    payload = build_payload(rows)
    pl, pg = payload["meta"]["totalLines"], round(payload["all"]["grandTotal"], 2)
    checks.append(("Dashboard payload reconciles to workbook", pl == N and abs(pg - GRAND) < 0.5,
                   f"{pl:,} lines / ${pg:,.2f}"))

    print("\n=== PASS CRITERIA ===")
    all_ok = True
    for name, ok, detail in checks:
        all_ok &= ok
        print(f"  [{'PASS' if ok else 'FAIL'}] {name:<46} ({detail})")

    # --- Informational: committed deliverables vs a fresh rebuild ---
    print("\n=== COMMITTED vs REBUILT (informational) ===")
    readme_c = load_grid(_paths.WORKBOOK, "README")
    rec_line = next((str(r[0]) for r in readme_c if r and str(r[0]).startswith("Records:")), "?")
    readme_ok = f"{N:,}" in rec_line
    any_drift = not readme_ok
    print(f"  README sheet: '{rec_line}'" + ("  (in sync)" if readme_ok else f"  <- All Lines holds {N:,}"))
    for sheet in ["By Year x Fund", "By Issue-Date FY x Fund", "By Budget Unit",
                  "PD_Yearly_Summary", "PD by Subject"]:
        rc = len(load_grid(_paths.WORKBOOK, sheet))
        rr = len(load_grid(SCRATCH, sheet))
        if rc != rr:
            any_drift = True
        print(f"  {sheet:<26} committed={rc:>5} | rebuilt={rr:>5}{'  <- differs' if rc != rr else ''}")
    if any_drift:
        print("  (Differences mean the committed workbook predates a clean rebuild — e.g. collapsed")
        print("   Fund 320/423 name variants or an older PD-sheet layout; run `rebuild.py --assemble-only`.)")

    html = _paths.DASHBOARD.read_text(encoding="utf-8")
    cp = json.loads(re.search(r'<script id="data-payload"[^>]*>(.*?)</script>', html, re.DOTALL).group(1))
    cl = cp["meta"]["totalLines"]
    print(f"  index.html payload: {cl:,} lines" +
          ("  (in sync)" if cl == N else f"  <- STALE by {N - cl:+,}; run `rebuild.py --assemble-only`"))

    SCRATCH.unlink(missing_ok=True)
    print("\n=== RESULT ===")
    print("PASS — consolidated pipeline reproduces the data faithfully and builds a consistent workbook"
          if all_ok else "FAIL — see failed criteria above")
    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
